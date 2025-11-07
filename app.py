# -*- coding: utf-8 -*-
# app.py
# App integrada: Coordenadas + Δh (ITM/MSAM) + HAAT + Predicción ITM (Simplificado)
# - Δh (ITM): 10–50 km, 500 m, Δh = h10 - h90, ΔF = 1.9 - 0.03*Δh*(1 + f/300)
# - HAAT: promedio terreno 3–16 km (500 m); HAAT = (elev_sitio + AGL_tx) - promedio_terreno
# - Predicción ITM Simplificado (estilo MSAM/FCC): Lp, E(dBµV/m) con ΔF y corrección de altura opcional
# - SRTM (srtm.py) online por defecto; ASTER/GTOPO30 (GeoTIFF EPSG:4326) opcional
# - Resultados persistentes, mapas Folium, Plotly y descargas

import streamlit as st
import pandas as pd
import numpy as np
import math
from io import BytesIO

from pygeodesy.ellipsoidalVincenty import LatLon
import folium
from streamlit_folium import st_folium
import plotly.graph_objects as go

import srtm
import rasterio

# ---------------------------
# Configuración App
# ---------------------------
st.set_page_config(page_title="Coordenadas + Δh + HAAT + ITM", layout="wide")
st.title("🧭 Coordenadas + 🌄 Δh (ITM / MSAM) + 📡 HAAT + 📶 Predicción ITM (Simplificado)")

# ---------------------------
# Estado persistente
# ---------------------------
if "categoria" not in st.session_state:
    st.session_state.categoria = "Cálculo - 8 Radiales"
if "resultados" not in st.session_state:
    st.session_state.resultados = {}
if "deltaH_state" not in st.session_state:
    st.session_state.deltaH_state = None
if "haat_state" not in st.session_state:
    st.session_state.haat_state = None
if "itm_state" not in st.session_state:
    st.session_state.itm_state = None

# ---------------------------
# Utilidades geodésicas / conversión
# ---------------------------
R_EARTH_M = 6371000.0

def destination_point(lat_deg, lon_deg, bearing_deg, distance_m):
    lat1 = math.radians(lat_deg)
    lon1 = math.radians(lon_deg)
    brng = math.radians(bearing_deg)
    dr = distance_m / R_EARTH_M
    lat2 = math.asin(math.sin(lat1) * math.cos(dr) + math.cos(lat1) * math.sin(dr) * math.cos(brng))
    lon2 = lon1 + math.atan2(math.sin(brng) * math.sin(dr) * math.cos(lat1),
                             math.cos(dr) - math.sin(lat1) * math.sin(lat2))
    return math.degrees(lat2), (math.degrees(lon2) + 540) % 360 - 180

def decimal_a_gms(grados_decimales, tipo):
    direccion = {"lat": "N" if grados_decimales >= 0 else "S",
                 "lon": "E" if grados_decimales >= 0 else "W"}[tipo]
    gabs = abs(grados_decimales)
    g = int(gabs)
    m_dec = (gabs - g) * 60
    m = int(m_dec)
    s = (m_dec - m) * 60
    return f"{g}° {m}' {s:.8f}\" {direccion}"

def gms_a_decimal(grados:int, minutos:int, segundos:float, direccion:str, tipo:str):
    if tipo == "lat" and not (0 <= abs(grados) <= 90):  raise ValueError("Latitud grados 0–90")
    if tipo == "lon" and not (0 <= abs(grados) <= 180): raise ValueError("Longitud grados 0–180")
    if not (0 <= minutos < 60):  raise ValueError("Minutos 0–59")
    if not (0 <= segundos < 60): raise ValueError("Segundos 0–59.999")
    if tipo == "lat" and direccion not in ("N","S"): raise ValueError("Dir lat N/S")
    if tipo == "lon" and direccion not in ("E","W"): raise ValueError("Dir lon E/W")
    dec = abs(grados) + minutos/60 + segundos/3600
    if direccion in ("S","W"): dec = -dec
    return dec

def input_decimal(label_lat, label_lon, key_prefix):
    c1, c2 = st.columns(2)
    with c1:
        lat_txt = st.text_input(label_lat, value="8.8066", key=f"{key_prefix}_lat_dec")
    with c2:
        lon_txt = st.text_input(label_lon, value="-82.5403", key=f"{key_prefix}_lon_dec")
    try:
        lat = float(lat_txt); lon = float(lon_txt)
    except ValueError:
        st.error("Lat/Lon decimales inválidos."); st.stop()
    st.caption(f"⇄ GMS: Lat {decimal_a_gms(lat,'lat')} | Lon {decimal_a_gms(lon,'lon')}")
    return lat, lon

def input_gms(key_prefix, defaults=("N","W")):
    st.write("**Latitud (GMS)**")
    a,b,c,d = st.columns([1,1,1,1])
    with a: lat_g = st.number_input("Grados", value=8, step=1, format="%d", key=f"{key_prefix}_lat_g")
    with b: lat_m = st.number_input("Min", value=48, min_value=0, max_value=59, step=1, format="%d", key=f"{key_prefix}_lat_m")
    with c: lat_s = st.number_input("Seg", value=23.76, min_value=0.0, max_value=59.999999, step=0.01, format="%.6f", key=f"{key_prefix}_lat_s")
    with d: lat_d = st.selectbox("Dir", ["N","S"], index=0 if defaults[0]=="N" else 1, key=f"{key_prefix}_lat_d")

    st.write("**Longitud (GMS)**")
    e,f,g,h = st.columns([1,1,1,1])
    with e: lon_g = st.number_input("Grados", value=82, step=1, format="%d", key=f"{key_prefix}_lon_g")
    with f: lon_m = st.number_input("Min", value=32, min_value=0, max_value=59, step=1, format="%d", key=f"{key_prefix}_lon_m")
    with g: lon_s = st.number_input("Seg", value=25.08, min_value=0.0, max_value=59.999999, step=0.01, format="%.6f", key=f"{key_prefix}_lon_s")
    with h: lon_d = st.selectbox("Dir", ["E","W"], index=1 if defaults[1]=="W" else 0, key=f"{key_prefix}_lon_d")

    try:
        lat = gms_a_decimal(lat_g, lat_m, lat_s, lat_d, "lat")
        lon = gms_a_decimal(lon_g, lon_m, lon_s, lon_d, "lon")
    except Exception as e:
        st.error(f"Error GMS: {e}"); st.stop()
    st.caption(f"⇄ Decimal: Lat {lat:.10f} | Lon {lon:.10f}")
    return lat, lon

def input_coords(key_prefix="base"):
    st.markdown("#### Formato de coordenadas de entrada")
    modo = st.radio("Formato", ["Decimal", "Grados, Minutos y Segundos (GMS)"], horizontal=True, key=f"{key_prefix}_fmt")
    if modo == "Decimal":
        return input_decimal("Latitud inicial (decimal)", "Longitud inicial (decimal)", key_prefix)
    return input_gms(key_prefix)

# ---------------------------
# DEM / Elevaciones
# ---------------------------
@st.cache_resource
def get_srtm_data():
    return srtm.get_data()  # descarga/caché automática (en línea por defecto)

@st.cache_resource
def open_raster(path_tif:str):
    try:
        return rasterio.open(path_tif)
    except Exception as e:
        st.warning(f"No se pudo abrir GeoTIFF: {e}")
        return None

def elev_srtm(lats, lons):
    data = get_srtm_data()
    return [data.get_elevation(la, lo) for la,lo in zip(lats,lons)]

def elev_raster(ds, lats, lons):
    if ds is None:
        return [None]*len(lats)
    band1 = ds.read(1)
    vals = []
    for la, lo in zip(lats, lons):
        try:
            row, col = ds.index(lo, la)  # (lon,lat) -> (row,col)
            v = band1[row, col]
            if ds.nodata is not None and v == ds.nodata:
                vals.append(None)
            else:
                vals.append(float(v))
        except Exception:
            vals.append(None)
    return vals

def get_site_elevation(lat, lon, use_srtm=True, ds_raster=None):
    return (elev_srtm([lat], [lon])[0] if use_srtm else elev_raster(ds_raster, [lat], [lon])[0])

# ---------------------------
# Cálculos de perfiles y métricas
# ---------------------------
def calcular_puntos(lat, lon, acimuts, distancias_m):
    base = LatLon(lat, lon)
    out = []
    for d in distancias_m:
        for az in acimuts:
            p = base.destination(d, az)
            out.append({
                "Distancia (km)": d/1000,
                "Acimut (°)": az,
                "Latitud Final (Decimal)": f"{p.lat:.10f}",
                "Longitud Final (Decimal)": f"{p.lon:.10f}",
                "Latitud (GMS)": decimal_a_gms(p.lat, "lat"),
                "Longitud (GMS)": decimal_a_gms(p.lon, "lon")
            })
    return pd.DataFrame(out)

def calcular_distancia_azimut(lat1, lon1, lat2, lon2):
    p1 = LatLon(lat1, lon1); p2 = LatLon(lat2, lon2)
    d = p1.distanceTo(p2) / 1000.0
    az12 = p1.initialBearingTo(p2); az21 = p2.initialBearingTo(p1)
    return d, az12, az21

def build_profile(lat0, lon0, az, start_km, end_km, step_m):
    dists_m = list(range(int(start_km*1000), int(end_km*1000)+1, int(step_m)))
    lats, lons = [], []
    for d in dists_m:
        la, lo = destination_point(lat0, lon0, az, d)
        lats.append(la); lons.append(lo)
    return dists_m, lats, lons

def compute_delta_h(elev_list):
    arr = np.array([e for e in elev_list if e is not None], dtype=float)
    if arr.size == 0:
        return None, None, None
    h10 = float(np.percentile(arr, 90))  # P90
    h90 = float(np.percentile(arr, 10))  # P10
    return h10 - h90, h10, h90

def avg_terrain(elev_list):
    arr = np.array([e for e in elev_list if e is not None], dtype=float)
    if arr.size == 0:
        return None
    return float(np.mean(arr))

def deltaF_from_deltaH(delta_h, f_mhz):
    return 1.9 - 0.03 * delta_h * (1 + f_mhz/300.0)

# Heurística de ganancia por altura (ligera, estilo práctico)
def height_gain_db(haat_m, h_rx_m):
    gh_tx = 0.0 if haat_m is None else 20.0 * math.log10(max(1.0, haat_m/30.0))
    gh_rx = 10.0 * math.log10(max(1.0, h_rx_m/10.0))
    return gh_tx + gh_rx

# ---------------------------
# Panel de categorías
# ---------------------------
st.markdown("### Selecciona la categoría de cálculo")
c1, c2 = st.columns(2); c3, c4 = st.columns(2); c5, c6 = st.columns(2); c7, _ = st.columns(2)

if c1.button("📍 Cálculo - 8 Radiales"):
    st.session_state.categoria = "Cálculo - 8 Radiales"
if c2.button("🧭 Cálculo por Azimut"):
    st.session_state.categoria = "Cálculo por Azimut"
if c3.button("📏 Cálculo de Distancia"):
    st.session_state.categoria = "Cálculo de Distancia"
if c4.button("🗺️ Cálculo de Distancia Central"):
    st.session_state.categoria = "Cálculo de Distancia Central"
if c5.button("🌄 Δh – Rugosidad (ITM)"):
    st.session_state.categoria = "Δh – Rugosidad (ITM)"
if c6.button("📡 Altura efectiva (HAAT)"):
    st.session_state.categoria = "Altura efectiva (HAAT)"
if c7.button("📶 Predicción ITM (Simplificado)"):
    st.session_state.categoria = "Predicción ITM (Simplificado)"

categoria = st.session_state.categoria
st.markdown(f"### 🟢 Categoría seleccionada: {categoria}")

# ---------------------------
# Coordenadas base
# ---------------------------
lat, lon = input_coords(key_prefix=f"{categoria}_base")

# ---------------------------
# Mapa genérico
# ---------------------------
def mostrar_mapa_generico(df, lat, lon, categoria):
    m = folium.Map(location=[lat, lon], zoom_start=9, control_scale=True)
    if categoria in ("Cálculo - 8 Radiales", "Cálculo por Azimut"):
        for _, r in df.iterrows():
            folium.Marker([float(r["Latitud Final (Decimal)"]),
                           float(r["Longitud Final (Decimal)"])],
                          tooltip=f"{r.get('Acimut (°)','')}° - {r.get('Distancia (km)','')} km").add_to(m)
        folium.Marker([lat, lon], tooltip="Punto inicial", icon=folium.Icon(color="red")).add_to(m)
    elif categoria == "Cálculo de Distancia":
        for _, r in df.iterrows():
            lat2, lon2 = float(r["Latitud 2"]), float(r["Longitud 2"])
            folium.Marker([lat, lon], tooltip="Punto 1", icon=folium.Icon(color="red")).add_to(m)
            folium.Marker([lat2, lon2], tooltip="Punto 2", icon=folium.Icon(color="blue")).add_to(m)
            folium.PolyLine([[lat, lon], [lat2, lon2]], weight=2).add_to(m)
    elif categoria == "Cálculo de Distancia Central":
        for _, r in df.iterrows():
            latc, lonc = float(r["Latitud central"]), float(r["Longitud central"])
            latp, lonp = float(r["Latitud punto"]), float(r["Longitud punto"])
            folium.Marker([latc, lonc], tooltip="Central", icon=folium.Icon(color="red")).add_to(m)
            folium.Marker([latp, lonp], tooltip="Punto", icon=folium.Icon(color="blue")).add_to(m)
            folium.PolyLine([[latc, lonc], [latp, lonp]], color="green", weight=2).add_to(m)
    st_folium(m, width=None, height=480)

# ---------------------------
# Pestañas de coordenadas estándar
# ---------------------------
if categoria == "Cálculo - 8 Radiales":
    acimuts = [0,45,90,135,180,225,270,315]
    dist_m = [10000, 50000]
    if st.button("Calcular", key="calc_8rad"):
        st.session_state.resultados[categoria] = calcular_puntos(lat, lon, acimuts, dist_m)

elif categoria == "Cálculo por Azimut":
    az_txt = st.text_input("Azimuts (°) separados por coma", value="0,45,90,135,180,225,270,315")
    d1 = st.number_input("Distancia 1 (m)", value=10000, min_value=1, step=100)
    d2 = st.number_input("Distancia 2 (m)", value=50000, min_value=1, step=100)
    if st.button("Calcular", key="calc_az"):
        try:
            acimuts = [float(a.strip()) for a in az_txt.split(",") if a.strip()!=""]
            st.session_state.resultados[categoria] = calcular_puntos(lat, lon, acimuts, [d1, d2])
        except Exception as e:
            st.error(f"Error en azimuts: {e}")

elif categoria == "Cálculo de Distancia":
    modo2 = st.radio("Formato para el Punto 2", ["Decimal","GMS"], horizontal=True, key="fmt_p2")
    if modo2 == "Decimal":
        c1, c2 = st.columns(2)
        with c1: lat2 = st.text_input("Latitud 2 (decimal)", value="8.8066")
        with c2: lon2 = st.text_input("Longitud 2 (decimal)", value="-82.5403")
        try:
            lat2f = float(lat2); lon2f = float(lon2)
        except ValueError:
            st.error("Lat/Lon decimales inválidos."); st.stop()
        st.caption(f"Punto 2 (GMS): Lat {decimal_a_gms(lat2f,'lat')} | Lon {decimal_a_gms(lon2f,'lon')}")
    else:
        lat2f, lon2f = input_gms(key_prefix="punto2", defaults=("N","W"))
    if st.button("Calcular", key="calc_dist"):
        dkm, az12, az21 = calcular_distancia_azimut(lat, lon, lat2f, lon2f)
        st.session_state.resultados[categoria] = pd.DataFrame([{
            "Distancia (km)": dkm,
            "Acimut ida (°)": az12,
            "Acimut vuelta (°)": az21,
            "Latitud 1": lat, "Longitud 1": lon,
            "Latitud 2": lat2f, "Longitud 2": lon2f
        }])

elif categoria == "Cálculo de Distancia Central":
    n = st.number_input("Número de puntos", min_value=1, value=2, step=1)
    filas = []
    for i in range(int(n)):
        modo_i = st.radio(f"Formato Punto {i+1}", ["Decimal","GMS"], horizontal=True, key=f"fmt_central_{i}")
        if modo_i == "Decimal":
            c1,c2 = st.columns(2)
            with c1: latp = st.text_input(f"Latitud punto {i+1} (decimal)", value="8.8066", key=f"latp_{i}")
            with c2: lonp = st.text_input(f"Longitud punto {i+1} (decimal)", value="-82.5403", key=f"lonp_{i}")
            try: latpf = float(latp); lonpf = float(lonp)
            except ValueError: st.error(f"Punto {i+1}: decimales inválidos."); st.stop()
            st.caption(f"Punto {i+1} (GMS): Lat {decimal_a_gms(latpf,'lat')} | Lon {decimal_a_gms(lonpf,'lon')}")
        else:
            latpf, lonpf = input_gms(key_prefix=f"punto{i+1}", defaults=("N","W"))
        dkm, az12, az21 = calcular_distancia_azimut(lat, lon, latpf, lonpf)
        filas.append({
            "Distancia (km)": dkm,
            "Acimut ida (°)": az12,
            "Acimut vuelta (°)": az21,
            "Latitud central": lat, "Longitud central": lon,
            "Latitud punto": latpf, "Longitud punto": lonpf
        })
    if st.button("Calcular", key="calc_central"):
        st.session_state.resultados[categoria] = pd.DataFrame(filas)

# ---------------------------
# Δh – Rugosidad (ITM) (MSAM)
# ---------------------------
if categoria == "Δh – Rugosidad (ITM)":
    st.markdown("#### Parámetros Δh (ITM/MSAM) — 10–50 km, paso 500 m")
    c = st.columns(5)
    with c[0]: fmhz = st.number_input("Frecuencia (MHz)", value=102.1, step=0.1, format="%.1f")
    with c[1]: az_txt = st.text_input("Azimuts (°) separados por coma", value="0,45,90,135,180,225,270,315")
    with c[2]: fuente = st.selectbox("Fuente elevación", ["SRTM (online por defecto)", "ASTER (GeoTIFF)", "GTOPO30 (GeoTIFF)"])
    with c[3]: aster_path = st.text_input("Ruta GeoTIFF (si ASTER/GTOPO30)", value="")
    with c[4]: ant_agl = st.number_input("Altura antena AGL (m) (para AMSL/HAAT)", value=50.0, min_value=0.0, step=1.0)

    if st.button("Calcular Δh (MSAM/ITM)", key="calc_dh_msam"):
        st.session_state.deltaH_state = {"status": "running"}
        try:
            az_list = [float(a.strip()) for a in az_txt.split(",") if a.strip()!=""]
        except:
            st.error("Revisa la lista de azimuts."); st.session_state.deltaH_state=None; st.stop()

        start_km, end_km, paso_m = 10.0, 50.0, 500
        ds_raster = None
        use_srtm = (fuente.startswith("SRTM"))
        if not use_srtm:
            if aster_path.strip():
                ds_raster = open_raster(aster_path.strip())
                if ds_raster is None:
                    st.info("No se pudo abrir el GeoTIFF. Se usará SRTM.")
                    use_srtm = True
            else:
                st.info("No se indicó ruta GeoTIFF. Se usará SRTM.")
                use_srtm = True

        elev_site = get_site_elevation(lat, lon, use_srtm, ds_raster)
        ant_amsl = (elev_site if elev_site is not None else 0.0) + ant_agl

        results = []; profiles = {}
        prog = st.progress(0); total = len(az_list)

        for i, az in enumerate(az_list, start=1):
            dists_m, lats, lons = build_profile(lat, lon, az, start_km, end_km, paso_m)
            elev = elev_srtm(lats, lons) if use_srtm else elev_raster(ds_raster, lats, lons)
            dh, h10, h90 = compute_delta_h(elev)
            row = {"Azimut (°)": az}
            if dh is not None:
                row["Δh (m)"] = round(dh, 2)
                row["ΔF (dB)"] = round(deltaF_from_deltaH(dh, fmhz), 2)
                row["Elevación sitio (m)"] = elev_site
                row["Antena AGL (m)"] = ant_agl
                row["Antena AMSL (m)"] = ant_amsl
                profiles[az] = pd.DataFrame({
                    "Distancia (km)": [d/1000 for d in dists_m],
                    "Elevación (m)": elev, "Lat": lats, "Lon": lons
                })
            results.append(row)
            prog.progress(int(i*100/total))

        res_df = pd.DataFrame(results).sort_values("Azimut (°)").reset_index(drop=True)
        st.session_state.deltaH_state = {
            "status": "done",
            "inputs": {"fmhz": fmhz, "azimuts": az_list, "fuente": fuente, "paso_m": paso_m, "AGL": ant_agl},
            "df": res_df,
            "profiles": profiles
        }

    if st.session_state.deltaH_state and st.session_state.deltaH_state.get("status") == "done":
        res_df = st.session_state.deltaH_state["df"]; profiles = st.session_state.deltaH_state["profiles"]
        st.subheader("Resultados Δh (MSAM/ITM) + alturas del sitio")
        st.dataframe(res_df, use_container_width=True)

        # Perfil
        az_opts = res_df["Azimut (°)"].tolist()
        if az_opts:
            az_sel = st.selectbox("Ver perfil (azimut):", az_opts)
            prof = profiles.get(az_sel)
            if prof is not None:
                fig = go.Figure()
                fig.add_trace(go.Scatter(x=prof["Distancia (km)"], y=prof["Elevación (m)"], mode="lines",
                                         name=f"Perfil – Az {az_sel}°"))
                fig.update_layout(title=f"Perfil 10–50 km – Az {az_sel}°", xaxis_title="Distancia (km)", yaxis_title="Elevación (m)")
                st.plotly_chart(fig, use_container_width=True)

        # Descargas
        def df_to_excel_bytes(df, sheet="DeltaH_ITM"):
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            wb = Workbook(); ws = wb.active; ws.title = sheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            ws["I1"] = "Δh = h10 - h90 (10–50 km, 500 m); ΔF = 1.9 - 0.03*Δh*(1 + f/300)"
            out = BytesIO(); wb.save(out); return out.getvalue()

        st.download_button("⬇️ CSV (Δh + alturas)", data=res_df.to_csv(index=False).encode("utf-8"),
                           file_name="deltaH_ITM_con_alturas.csv", mime="text/csv")
        st.download_button("⬇️ Excel (Δh + alturas)", data=df_to_excel_bytes(res_df),
                           file_name="deltaH_ITM_con_alturas.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ---------------------------
# 📡 Altura efectiva (HAAT)
# ---------------------------
if categoria == "Altura efectiva (HAAT)":
    st.markdown("#### Parámetros HAAT — promedio terreno 3–16 km (paso 500 m)")
    c = st.columns(5)
    with c[0]: az_txt_h = st.text_input("Azimuts (°) separados por coma", value="0,45,90,135,180,225,270,315")
    with c[1]: fuente_h = st.selectbox("Fuente elevación", ["SRTM (online por defecto)", "ASTER (GeoTIFF)", "GTOPO30 (GeoTIFF)"], key="fuente_haat")
    with c[2]: aster_path_h = st.text_input("Ruta GeoTIFF (si ASTER/GTOPO30)", value="", key="tif_haat")
    with c[3]: ant_agl_h = st.number_input("Altura antena AGL (m)", value=50.0, min_value=0.0, step=1.0, key="agl_haat")
    with c[4]: incluir_perfil = st.checkbox("Exportar perfiles 3–16 km", value=True)

    if st.button("Calcular HAAT por azimut", key="calc_haat_btn"):
        st.session_state.haat_state = {"status": "running"}
        try:
            az_list_h = [float(a.strip()) for a in az_txt_h.split(",") if a.strip()!=""]
        except:
            st.error("Revisa la lista de azimuts."); st.session_state.haat_state=None; st.stop()

        start_km, end_km, paso_m = 3.0, 16.0, 500
        ds_raster = None
        use_srtm = (fuente_h.startswith("SRTM"))
        if not use_srtm:
            if aster_path_h.strip():
                ds_raster = open_raster(aster_path_h.strip())
                if ds_raster is None:
                    st.info("No se pudo abrir el GeoTIFF. Se usará SRTM.")
                    use_srtm = True
            else:
                st.info("No se indicó ruta GeoTIFF. Se usará SRTM.")
                use_srtm = True

        elev_site = get_site_elevation(lat, lon, use_srtm, ds_raster)
        ant_amsl = (elev_site if elev_site is not None else 0.0) + ant_agl_h

        results = []; profiles = {}
        prog = st.progress(0); total = len(az_list_h)

        for i, az in enumerate(az_list_h, start=1):
            dists_m, lats, lons = build_profile(lat, lon, az, start_km, end_km, paso_m)
            elev = elev_srtm(lats, lons) if use_srtm else elev_raster(ds_raster, lats, lons)
            avg_terr = avg_terrain(elev)
            row = {"Azimut (°)": az, "Elevación sitio (m)": elev_site, "Antena AGL (m)": ant_agl_h, "Antena AMSL (m)": ant_amsl}
            if avg_terr is not None:
                row["Promedio terreno 3–16 km (m)"] = round(avg_terr,2)
                row["HAAT (m)"] = round(ant_amsl - avg_terr, 2)
            results.append(row)

            if incluir_perfil:
                profiles[az] = pd.DataFrame({
                    "Distancia (km)": [d/1000 for d in dists_m],
                    "Elevación (m)": elev, "Lat": lats, "Lon": lons
                })
            prog.progress(int(i*100/total))

        res_df = pd.DataFrame(results).sort_values("Azimut (°)").reset_index(drop=True)
        st.session_state.haat_state = {"status":"done","df":res_df,"profiles":profiles}

    if st.session_state.haat_state and st.session_state.haat_state.get("status") == "done":
        res_df = st.session_state.haat_state["df"]; profiles = st.session_state.haat_state["profiles"]
        st.subheader("Resultados HAAT por azimut")
        st.dataframe(res_df, use_container_width=True)

        if "HAAT (m)" in res_df.columns and not res_df["HAAT (m)"].dropna().empty:
            st.markdown("**Resumen:**")
            st.write({
                "HAAT (m) promedio": round(res_df["HAAT (m)"].dropna().mean(), 2),
                "Terreno 3–16 km promedio (m)": round(res_df["Promedio terreno 3–16 km (m)"].dropna().mean(), 2) if "Promedio terreno 3–16 km (m)" in res_df.columns else None
            })

        # Descargas
        def df_to_excel_bytes(df, sheet="HAAT"):
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            wb = Workbook(); ws = wb.active; ws.title = sheet
            for r in dataframe_to_rows(df, index=False, header=True): ws.append(r)
            ws["I1"] = "HAAT = (Elevación sitio + AGL_tx) - promedio terreno 3–16 km"
            out = BytesIO(); wb.save(out); return out.getvalue()

        st.download_button("⬇️ CSV (HAAT)", data=res_df.to_csv(index=False).encode("utf-8"),
                           file_name="haat_por_azimut.csv", mime="text/csv")
        st.download_button("⬇️ Excel (HAAT)", data=df_to_excel_bytes(res_df, "HAAT"),
                           file_name="haat_por_azimut.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ---------------------------
# 📶 Predicción ITM (Simplificado)
# ---------------------------
if categoria == "Predicción ITM (Simplificado)":
    st.markdown("#### Entradas de Predicción (estilo MSAM/FCC)")

    col = st.columns(6)
    with col[0]: fmhz_p = st.number_input("Frecuencia (MHz)", value=100.0, min_value=30.0, step=0.1, format="%.1f")
    with col[1]: erp_kw = st.number_input("ERP (kW)", value=5.0, min_value=0.001, step=0.1, format="%.3f")
    with col[2]: az_p = st.number_input("Azimut (°)", value=0.0, min_value=0.0, max_value=359.9, step=0.1)
    with col[3]: h_tx_agl = st.number_input("Altura TX AGL (m)", value=50.0, min_value=0.0, step=1.0)
    with col[4]: h_rx_agl = st.number_input("Altura RX AGL (m)", value=10.0, min_value=0.0, step=0.5)
    with col[5]: aplicar_hg = st.checkbox("Aplicar corrección de altura (G_h)", value=True)

    col2 = st.columns(4)
    with col2[0]: fuente_p = st.selectbox("Fuente elevación", ["SRTM (online por defecto)", "ASTER (GeoTIFF)", "GTOPO30 (GeoTIFF)"])
    with col2[1]: aster_path_p = st.text_input("Ruta GeoTIFF (si ASTER/GTOPO30)", value="")
    with col2[2]: d_km_max = st.number_input("Distancia máx (km)", value=60.0, min_value=1.0, step=1.0, format="%.1f")
    with col2[3]: paso_km = st.number_input("Paso (km)", value=1.0, min_value=0.1, step=0.1, format="%.1f")

    st.caption("Notas: Δh se evalúa 10–50 km; HAAT a 3–16 km. Si no hay resultados previos, la app los calcula automáticamente para el azimut indicado.")

    if st.button("Calcular Predicción ITM (Simplificado)", key="calc_itm_btn"):
        st.session_state.itm_state = {"status":"running"}

        # Fuente DEM
        ds_raster = None
        use_srtm = (fuente_p.startswith("SRTM"))
        if not use_srtm:
            if aster_path_p.strip():
                ds_raster = open_raster(aster_path_p.strip())
                if ds_raster is None:
                    st.info("No se pudo abrir el GeoTIFF. Se usará SRTM.")
                    use_srtm = True
            else:
                st.info("No se indicó ruta GeoTIFF. Se usará SRTM.")
                use_srtm = True

        # Elevación sitio
        elev_site = get_site_elevation(lat, lon, use_srtm, ds_raster)
        ant_amsl = (elev_site if elev_site is not None else 0.0) + h_tx_agl

        # 1) Δh (10–50 km, 500 m) — usar de estado previo si existe, si no calcular para este azimut
        delta_h_val = None
        if st.session_state.deltaH_state and st.session_state.deltaH_state.get("status")=="done":
            df_dh = st.session_state.deltaH_state["df"]
            row = df_dh.loc[df_dh["Azimut (°)"].round(1) == round(az_p,1)]
            if not row.empty and "Δh (m)" in row.columns:
                delta_h_val = float(row.iloc[0]["Δh (m)"])
        if delta_h_val is None:
            # calcular on-the-fly
            dists_m, lats, lons = build_profile(lat, lon, az_p, 10.0, 50.0, 500)
            elevs = elev_srtm(lats, lons) if use_srtm else elev_raster(ds_raster, lats, lons)
            delta_h_val, _, _ = compute_delta_h(elevs)

        # 2) HAAT (3–16 km, 500 m) — usar de estado previo si existe, si no calcular
        haat_val = None
        if st.session_state.haat_state and st.session_state.haat_state.get("status")=="done":
            df_h = st.session_state.haat_state["df"]
            rowh = df_h.loc[df_h["Azimut (°)"].round(1) == round(az_p,1)]
            if not rowh.empty and "HAAT (m)" in rowh.columns:
                haat_val = float(rowh.iloc[0]["HAAT (m)"])
        if haat_val is None:
            d2_m, lats2, lons2 = build_profile(lat, lon, az_p, 3.0, 16.0, 500)
            elevs2 = elev_srtm(lats2, lons2) if use_srtm else elev_raster(ds_raster, lats2, lons2)
            avgterr = avg_terrain(elevs2)
            haat_val = (ant_amsl - avgterr) if (avgterr is not None) else None

        # 3) Calcular curva de pérdida/ campo vs distancia
        d_vals = np.arange(1.0, d_km_max+1e-9, paso_km, dtype=float)
        erp_dBk = 10.0 * math.log10(erp_kw)  # ERP en kW -> dBk
        series = []
        for d_km in d_vals:
            fspl = 32.45 + 20.0*math.log10(fmhz_p) + 20.0*math.log10(max(d_km, 1e-6))
            dF = 0.0
            if delta_h_val is not None:
                dF = deltaF_from_deltaH(delta_h_val, fmhz_p)  # puede ser negativo -> reduce pérdida
            Gh = height_gain_db(haat_val, h_rx_agl) if aplicar_hg else 0.0
            Lp = fspl - dF - Gh
            E = 106.92 + erp_dBk - Lp  # dBµV/m
            series.append({"Distancia (km)": d_km, "Lp (dB)": Lp, "E (dBµV/m)": E})

        pred_df = pd.DataFrame(series)
        st.session_state.itm_state = {
            "status":"done",
            "inputs":{"fmhz": fmhz_p, "erp_kw": erp_kw, "az": az_p, "h_tx_agl": h_tx_agl, "h_rx_agl": h_rx_agl,
                      "delta_h": delta_h_val, "haat": haat_val, "aplicar_hg": aplicar_hg},
            "df": pred_df
        }

    # Mostrar resultados persistentes de la predicción
    if st.session_state.itm_state and st.session_state.itm_state.get("status")=="done":
        info = st.session_state.itm_state["inputs"]
        pred_df = st.session_state.itm_state["df"]

        st.subheader("Resultados de Predicción (ITM simplificado)")
        cols = st.columns(5)
        cols[0].metric("Δh usado (m)", f"{info.get('delta_h', None):.2f}" if info.get('delta_h') else "—")
        cols[1].metric("HAAT usado (m)", f"{info.get('haat', None):.2f}" if info.get('haat') else "—")
        cols[2].metric("ERP (dBk)", f"{10*math.log10(info['erp_kw']):.2f}")
        cols[3].metric("f (MHz)", f"{info['fmhz']:.1f}")
        cols[4].metric("Azimut (°)", f"{info['az']:.1f}")

        st.dataframe(pred_df, use_container_width=True)

        # Gráfico
        fig = go.Figure()
        fig.add_trace(go.Scatter(x=pred_df["Distancia (km)"], y=pred_df["E (dBµV/m)"],
                                 mode="lines", name="Campo E"))
        fig.add_trace(go.Scatter(x=pred_df["Distancia (km)"], y=pred_df["Lp (dB)"],
                                 mode="lines", name="Pérdida Lp", yaxis="y2"))
        fig.update_layout(
            title="Campo y Pérdida vs Distancia",
            xaxis_title="Distancia (km)",
            yaxis=dict(title="E (dBµV/m)"),
            yaxis2=dict(title="Lp (dB)", overlaying="y", side="right")
        )
        st.plotly_chart(fig, use_container_width=True)

        # Descargas
        def df_to_excel_bytes(df, sheet="Pred_ITM"):
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            wb = Workbook(); ws = wb.active; ws.title = sheet
            for r in dataframe_to_rows(df, index=False, header=True): ws.append(r)
            ws["G1"] = "Lp = 32.45 + 20log f + 20log d - ΔF - G_h ; E(dBµV/m) = 106.92 + ERP(dBk) - Lp"
            out = BytesIO(); wb.save(out); return out.getvalue()

        st.download_button("⬇️ CSV (Predicción ITM)", data=pred_df.to_csv(index=False).encode("utf-8"),
                           file_name="prediccion_itm_simplificado.csv", mime="text/csv")
        st.download_button("⬇️ Excel (Predicción ITM)", data=df_to_excel_bytes(pred_df),
                           file_name="prediccion_itm_simplificado.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ---------------------------
# Mostrar resultados persistentes de otras categorías
# ---------------------------
if categoria in st.session_state.resultados and categoria not in ("Δh – Rugosidad (ITM)", "Altura efectiva (HAAT)", "Predicción ITM (Simplificado)"):
    df = st.session_state.resultados[categoria]
    st.subheader("Resultados")
    if "Distancia (km)" in df.columns and categoria in ("Cálculo - 8 Radiales", "Cálculo por Azimut"):
        for d in sorted(df["Distancia (km)"].unique()):
            st.markdown(f"**Resultados a {d} km**")
            st.dataframe(df[df["Distancia (km)"] == d], use_container_width=True)
    else:
        st.dataframe(df, use_container_width=True)
    mostrar_mapa_generico(df, lat, lon, categoria)
    st.download_button(
        "📥 Descargar CSV",
        data=df.to_csv(index=False, sep=';', encoding='utf-8'),
        file_name=f"{categoria.replace(' ','_')}.csv",
        mime="text/csv"
    )
